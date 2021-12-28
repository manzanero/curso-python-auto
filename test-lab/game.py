from ursina import *
from openpyxl import load_workbook

from player import Player


app = Ursina()

sky = Sky(texture='sky_default')

text = Text(text="escape = quit", color=color.black, scale=(2, 2), x=-0.45, y=0.45, parent=camera.ui)


class LevelGrid(object):

    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.wb = load_workbook(excel_file, data_only=True)
        self.ws = self.wb.worksheets[0]
        self.max_row = self.ws.max_row
        self.max_col = self.ws.max_column

    def get_cell_type(self, col: int, row: int) -> str:
        try:
            color_hex = self.ws.cell(row, col).fill.start_color.index  # color hexadecimal
        except ValueError:
            return 'void'

        color_rgb = int(color_hex[2:4], 16), int(color_hex[4:6], 16), int(color_hex[6:8], 16)  # color RGB
        category = {
            (0, 0, 0): 'void',
            (146, 208, 80): 'start',
            (255, 0, 0): 'end',
            (0, 176, 80): 'grass',
            (192, 0, 0): 'wall',
            (255, 255, 0): 'door',
            (255, 192, 0): 'inner',
        }

        try:
            return category[color_rgb]
        except KeyError:
            raise Exception(f"Not a valid color: ({col}, {row})")


grid = LevelGrid('level.xlsx')

x_max = grid.max_col + 2
z_min = -(grid.max_row + 2)

for z in range(0, z_min, -1):
    for x in range(x_max):
        cell_type = grid.get_cell_type(x, -z)
        if cell_type == 'start':
            start = Entity(model="cube", scale=(1, 1, 1), color=color.rgba(146, 208, 80),
                           collider="box", texture="white_cube", position=(x, 0, z), rotation=(0, 0, 0))
            player = Player((x, 1.5, z), (0, 180, 0))

        elif cell_type == 'end':
            end = Entity(model="cube", scale=(1, 1, 1), color=color.rgba(255, 0, 0),
                         collider="box", texture="white_cube", position=(x, 0, z), rotation=(0, 0, 0))

        elif cell_type == 'wall':
            Entity(model="cube", scale=(1, 5, 1), color=color.rgba(255, 255, 255),
                   collider="box", texture="brick", texture_scale=(1, 5), position=(x, 2, z), rotation=(0, 0, 0))
            Entity(model="cube", scale=(1, 1, 1), color=color.rgba(192, 0, 0),
                   collider="box", texture="brick", texture_scale=(2, 2), position=(x, 5, z), rotation=(0, 0, 0))

        elif cell_type == 'inner':
            Entity(model="cube", scale=(1, 1, 1), color=color.rgba(255, 192, 0),
                   collider="box", texture="brick", texture_scale=(0.5, 0.5), position=(x, 0, z), rotation=(0, 0, 0))
            Entity(model="cube", scale=(1, 1, 1), color=color.rgba(192, 0, 0),
                   collider="box", texture="brick", texture_scale=(2, 2), position=(x, 5, z), rotation=(0, 0, 0))

        elif cell_type == 'door':
            Entity(model="cube", scale=(1, 1, 1), color=color.rgba(255, 192, 0),
                   collider="box", texture="brick", texture_scale=(0.5, 0.5), position=(x, 0, z), rotation=(0, 0, 0))
            Entity(model="cube", scale=(1, 1, 1), color=color.rgba(255, 255, 255),
                   collider="box", texture="brick", position=(x, 4, z), rotation=(0, 0, 0))
            Entity(model="cube", scale=(1, 1, 1), color=color.rgba(192, 0, 0),
                   collider="box", texture="brick", texture_scale=(2, 2), position=(x, 5, z), rotation=(0, 0, 0))

        elif cell_type == 'grass':
            Entity(model="cube", scale=(1, 1, 1), color=color.rgba(0, 176, 80),
                   texture_scale=(0.1, 0.1),
                   collider="box", texture="grass", position=(x, 0, z), rotation=(0, 0, 0))

        else:
            Entity(model="cube", scale=(1, 10, 1), color=color.rgba(0, 0, 0, 0),
                   collider="box", texture="white_cube", position=(x, 0, z), rotation=(0, 0, 0))


def update():
    ray = raycast(player.position, player.down, distance=2, ignore=[player, ])  # NoQA

    if ray.entity == end and player.enabled:
        mouse.locked = False
        player.disable()

        Text(text="The End", color=color.black, scale=(5, 5), origin=(0, 0), y=0.2, parent=camera.ui)

        btn__quit = Button(text="Quit", color=color.black, scale_y=0.1, scale_x=0.3, y=-0.2, parent=camera.ui)
        btn__quit.on_click = application.quit


class PauseMenu(Entity):

    def __init__(self, player):
        super().__init__(parent=camera.ui)
        self.player = player

        btn__resume = Button(text="Resume", color=color.black, scale_y=0.1, scale_x=0.3, y=0, parent=self)
        btn__resume.on_click = self.resume

        btn__quit = Button(text="Quit", color=color.black, scale_y=0.1, scale_x=0.3, y=-0.2, parent=self)
        btn__quit.on_click = application.quit

    def resume(self):
        self.player.enable()
        mouse.locked = True
        destroy(self)


def input(key):  # NoQA
    if key == "escape" and player.enabled:
        mouse.locked = False
        player.disable()
        PauseMenu(player)


app.run()
